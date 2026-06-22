#!/usr/bin/env python3
"""从 CF动作库 xlsx（动作库 sheet）生成 WODTrack/Config/SkillData.swift。

字段会持续演进时，更新 xlsx 后重跑本脚本即可重新生成 Swift 数据。
用法: python3 scripts/gen_skilldata.py <xlsx路径>
"""
import json
import sys
from collections import OrderedDict

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    "/Users/chen/Downloads/CF动作库_Keep_进阶链.xlsx"
OUT = "WODTrack/Config/SkillData.swift"

CAT_ID = {"体操": "gymnastics", "举重": "weightlifting"}
CAT_PREFIX = {"gymnastics": "gym", "weightlifting": "wl"}
TIER = {1: ".l1", 2: ".l2", 3: ".l3", 4: ".l4", 5: ".l5"}

# xlsx「动作库」列索引（0-based，对应 row[i]）。
# 表头：动作编号 | 动作名称 | 动作名称_En | 训练类别 | 动作模式 | 计量方式 | 难度 |
#       动作简介 | 动作简介_En | 动作步骤 | 动作步骤_En | 要点与注意 | 要点与注意_En |
#       前置动作 | 前置依据 | 降阶替代 | 源ID | 英文名称 | 视频链接
# 中文列后紧跟对应 _En 列；「降阶替代」无英文列 → englishScaling 留空，运行时回退中文。
ID, NAME, EN_NAME = 0, 1, 2
CATEGORY, PATTERN, MEASURE, TIER_COL = 3, 4, 5, 6
INTRO, EN_INTRO = 7, 8
STEPS, EN_STEPS = 9, 10
KEYPOINTS, EN_KEYPOINTS = 11, 12
PREREQS, RATIONALE, SCALING = 13, 14, 15
EN_SCALING = 16  # 降阶替代_En


def cell(r, i):
    """安全取列：越界返回 None。"""
    return r[i] if i < len(r) else None


def q(v):
    """转成 Swift 字符串字面量；空/None → nil。"""
    if v is None or str(v).strip() == "":
        return "nil"
    return json.dumps(str(v).strip(), ensure_ascii=False)


def qs(v):
    """必填字符串（空也给空串而非 nil）。"""
    s = "" if v is None else str(v).strip()
    return json.dumps(s, ensure_ascii=False)


def split_prereqs(v):
    if not v:
        return []
    import re
    return [t for t in re.split(r"[；;，,、\s]+", str(v).strip()) if t]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["动作库"]
    # 读全部列（含交错的 _En 英文列）。
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
    rows = [r for r in rows if r[ID]]

    # 每个类别按出现顺序收集 distinct 动作模式 → 子类 id
    sub_ids = {}   # (catId, pattern) -> subId
    cat_subs = {cid: OrderedDict() for cid in CAT_ID.values()}
    for r in rows:
        cid = CAT_ID[str(r[CATEGORY]).strip()]
        pattern = str(r[PATTERN]).strip()
        if pattern not in cat_subs[cid]:
            idx = len(cat_subs[cid]) + 1
            sid = f"{CAT_PREFIX[cid]}_m{idx:02d}"
            cat_subs[cid][pattern] = sid
        sub_ids[(cid, pattern)] = cat_subs[cid][pattern]

    def emit_skills(cid):
        out = []
        for r in rows:
            if CAT_ID[str(r[CATEGORY]).strip()] != cid:
                continue
            sid = sub_ids[(cid, str(r[PATTERN]).strip())]
            prereqs = split_prereqs(r[PREREQS])
            pre = "[" + ", ".join(json.dumps(p) for p in prereqs) + "]"
            tier = TIER[int(float(r[TIER_COL]))]
            # 5th/6th 实参顺序：动作模式 → measure 槽、计量 → pattern 槽（沿用既有生成约定，保持输出一致）。
            out.append(
                f'        s({q(r[ID])}, {qs(r[NAME])}, {q(sid)}, {tier}, {qs(r[PATTERN])}, {qs(r[MEASURE])},\n'
                f'          intro: {qs(r[INTRO])}, steps: {qs(r[STEPS])}, keyPoints: {q(r[KEYPOINTS])},\n'
                f'          scaling: {qs(r[SCALING])}, prereqs: {pre}, rationale: {q(r[RATIONALE])}, englishName: {qs(r[EN_NAME])},\n'
                f'          englishIntro: {qs(cell(r, EN_INTRO))}, englishSteps: {qs(cell(r, EN_STEPS))},\n'
                f'          englishKeyPoints: {q(cell(r, EN_KEYPOINTS))}, englishScaling: {qs(cell(r, EN_SCALING) if EN_SCALING is not None else None)})'
            )
        return ",\n".join(out)

    def emit_subs(cid):
        out = []
        for pattern, sid in cat_subs[cid].items():
            out.append(f'            SkillSubcategoryDefinition(id: {q(sid)}, name: {qs(pattern)}, categoryId: {q(cid)}),')
        return "\n".join(out)

    gym_total = sum(1 for r in rows if CAT_ID[str(r[CATEGORY]).strip()] == "gymnastics")
    wl_total = sum(1 for r in rows if CAT_ID[str(r[CATEGORY]).strip()] == "weightlifting")

    swift = f'''import Foundation
import SwiftUI

// ⚠️ 本文件由 scripts/gen_skilldata.py 从 CF动作库 xlsx 生成，请勿手改；改数据请更新 xlsx 后重跑脚本。

// MARK: - SkillTier

enum SkillTier: String, CaseIterable, Identifiable, Codable {{
    case l1, l2, l3, l4, l5
    var id: String {{ rawValue }}

    var label: String {{
        switch self {{
        case .l1: "L1"
        case .l2: "L2"
        case .l3: "L3"
        case .l4: "L4"
        case .l5: "L5"
        }}
    }}

    var description: String {{
        switch self {{
        case .l1: "入门"
        case .l2: "基础"
        case .l3: "进阶"
        case .l4: "高阶"
        case .l5: "精英"
        }}
    }}
}}

// MARK: - SkillSubcategoryDefinition

struct SkillSubcategoryDefinition: Identifiable {{
    let id: String
    let name: String
    let categoryId: String
}}

// MARK: - SkillDefinition

struct SkillDefinition: Identifiable {{
    let id: String
    let name: String
    /// 英文名称
    let englishName: String
    let tier: SkillTier
    let categoryId: String
    let subcategoryId: String
    /// 动作模式（= 子类名称，如「核心」「上肢拉(垂直)」）
    let movementPattern: String
    /// 计量方式（次数 / 静力时长 / 距离 …）
    let measure: String
    /// 动作简介
    let intro: String
    /// 动作步骤
    let steps: String
    /// 要点与注意（部分动作有）
    let keyPoints: String?
    /// 降阶替代
    let scaling: String
    /// 前置动作 id
    let prerequisites: [String]
    /// 前置依据
    let prereqRationale: String?
    /// 英文简介 / 步骤 / 要点 / 降阶（来自 xlsx 英文列；空则运行时回退中文）
    let englishIntro: String
    let englishSteps: String
    let englishKeyPoints: String?
    let englishScaling: String
}}

// MARK: - SkillCategoryDefinition

struct SkillCategoryDefinition: Identifiable {{
    let id: String
    let name: String
    let englishName: String
    let icon: String
    let color: Color
    let subcategories: [SkillSubcategoryDefinition]
    let skills: [SkillDefinition]
}}

// MARK: - SkillLibrary

enum SkillLibrary {{
    static let categories: [SkillCategoryDefinition] = [gymnastics, weightlifting]

    static var allSkills: [SkillDefinition] {{ categories.flatMap {{ $0.skills }} }}
    static var skillById: [String: SkillDefinition] {{
        Dictionary(uniqueKeysWithValues: allSkills.map {{ ($0.id, $0) }})
    }}

    /// 已上传示范视频的动作 id（在脚本模板内维护，重跑脚本不丢失）。
    /// 上传新视频到 COS 的 `movements/{{id}}.mp4` 后，把对应 id 加进来即可生效。
    static let skillsWithVideo: Set<String> = ["row"]

    static func hasVideo(_ skillId: String) -> Bool {{ skillsWithVideo.contains(skillId) }}

    // MARK: 体操 Gymnastics ({gym_total})

    static let gymnastics = SkillCategoryDefinition(
        id: "gymnastics",
        name: "体操",
        englishName: "Gymnastics",
        icon: "figure.gymnastics",
        color: Color(hex: "#5BE88F"),
        subcategories: [
{emit_subs("gymnastics")}
        ],
        skills: gymnasticSkills
    )

    private static let gymnasticSkills: [SkillDefinition] = [
{emit_skills("gymnastics")}
    ]

    // MARK: 举重 Weightlifting ({wl_total})

    static let weightlifting = SkillCategoryDefinition(
        id: "weightlifting",
        name: "举重",
        englishName: "Weightlifting",
        icon: "figure.strengthtraining.traditional",
        color: Color(hex: "#5BE88F"),
        subcategories: [
{emit_subs("weightlifting")}
        ],
        skills: weightliftingSkills
    )

    private static let weightliftingSkills: [SkillDefinition] = [
{emit_skills("weightlifting")}
    ]

    // MARK: - Helper

    private static func s(_ id: String, _ name: String, _ sub: String, _ tier: SkillTier,
                          _ measure: String, _ pattern: String,
                          intro: String, steps: String, keyPoints: String?,
                          scaling: String, prereqs: [String] = [], rationale: String? = nil,
                          englishName: String = "",
                          englishIntro: String = "", englishSteps: String = "",
                          englishKeyPoints: String? = nil, englishScaling: String = "") -> SkillDefinition {{
        // 注意：参数顺序 measure 在 pattern 之前是生成器约定；此处显式映射，避免歧义。
        SkillDefinition(
            id: id, name: name, englishName: englishName, tier: tier,
            categoryId: sub.hasPrefix("gym") ? "gymnastics" : "weightlifting",
            subcategoryId: sub, movementPattern: pattern, measure: measure,
            intro: intro, steps: steps, keyPoints: keyPoints,
            scaling: scaling, prerequisites: prereqs, prereqRationale: rationale,
            englishIntro: englishIntro, englishSteps: englishSteps,
            englishKeyPoints: englishKeyPoints, englishScaling: englishScaling
        )
    }}
}}
'''
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(swift)
    print(f"wrote {OUT}: gym={gym_total} wl={wl_total} total={gym_total + wl_total}")
    print("gym subs:", list(cat_subs["gymnastics"].items()))
    print("wl subs:", list(cat_subs["weightlifting"].items()))


if __name__ == "__main__":
    main()
