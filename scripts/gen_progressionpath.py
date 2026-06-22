#!/usr/bin/env python3
"""从 CF解锁路径 xlsx 生成 WODTrack/Config/ProgressionPathData.swift。

xlsx 列：路径ID | 动作 | 动作ID | 路径动作 | 路径动作ID | 路径说明
- 路径ID        → id
- 动作          → name（路线标题，即代表/目标动作的中文名）
- 动作ID        → targetSkillId（代表/目标动作，仅作身份，不并入站点）
- 路径动作ID    → steps（按 ; 切分，metro 站点序列，权威顺序）
- 路径说明      → intro（详情页顶部说明）
icon 不在 xlsx 中，由下方 ICON 查表给出。

用法: python3 scripts/gen_progressionpath.py <xlsx路径>
"""
import json
import re
import sys

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    "/Users/chen/Downloads/CF解锁路径 Final.xlsx"
OUT = "WODTrack/Config/ProgressionPathData.swift"

# path_id → SF Symbol（沿用技能树既有图标风格）
ICON = {
    "path_1": "circle.dashed",
    "path_2": "figure.handball",
    "path_3": "arrow.up.to.line.circle.fill",
    "path_4": "figure.strengthtraining.functional",
    "path_5": "flame.fill",
}
DEFAULT_ICON = "point.topleft.down.to.point.bottomright.curvepath.fill"

# path_id → 路线英文名（xlsx 无英文名列，与 ICON 同样在脚本内查表维护）。
EN_NAME = {
    "path_1": "Jump Rope",
    "path_2": "Handstand",
    "path_3": "Bar Muscle-Up",
    "path_4": "Pull-Up",
    "path_5": "Toes-to-Bar",
}


def q(v):
    s = "" if v is None else str(v).strip()
    return json.dumps(s, ensure_ascii=False)


def split_steps(v):
    if not v:
        return []
    return [t for t in re.split(r"[；;，,、\s]+", str(v).strip()) if t]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    # 列：路径ID | 动作(name) | 动作ID(target) | 路径动作 | 路径动作ID(steps) | 路径说明(intro) | 路径说明_En
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
    rows = [r for r in rows if r[0]]

    def cell(r, i):
        return r[i] if i < len(r) else None

    blocks = []
    for r in rows:
        pid, name, target, _names, step_ids, intro = r[:6]
        pid = str(pid).strip()
        en_name = EN_NAME.get(pid, "")          # 路线英文名走查表
        en_intro = cell(r, 6)                   # 路径说明_En（第7列）
        steps = split_steps(step_ids)
        steps_lit = "[\n" + "".join(f"                {q(s)},\n" for s in steps) + "            ]"
        blocks.append(
            f'        ProgressionPath(\n'
            f'            id: {q(pid)},\n'
            f'            name: {q(name)},\n'
            f'            englishName: {q(en_name)},\n'
            f'            icon: {q(ICON.get(pid, DEFAULT_ICON))},\n'
            f'            targetSkillId: {q(target)},\n'
            f'            intro: {q(intro)},\n'
            f'            englishIntro: {q(en_intro)},\n'
            f'            steps: {steps_lit}\n'
            f'        )'
        )

    body = ",\n".join(blocks)
    out = f'''import Foundation
import SwiftUI

// ⚠️ 本文件由 scripts/gen_progressionpath.py 从 CF解锁路径 xlsx 生成，请勿手改；改数据请更新 xlsx 后重跑脚本。

// MARK: - ProgressionPath

struct ProgressionPath: Identifiable {{
    let id: String
    let name: String
    /// 英文路线名（来自 xlsx；空则运行时回退中文）。
    var englishName: String = ""
    let icon: String
    /// 代表/目标动作（动作ID），仅作路线身份，不作为站点。
    let targetSkillId: String
    /// 路径说明，展示在路线详情页顶部。
    let intro: String
    /// 英文路径说明（来自 xlsx；空则运行时回退中文）。
    var englishIntro: String = ""
    /// 需按顺序依次解锁的动作 id 序列（metro 站点）。
    let steps: [String]
}}

// MARK: - ProgressionPathLibrary

enum ProgressionPathLibrary {{
    static let all: [ProgressionPath] = [
{body}
    ]
}}
'''
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(out)
    print(f"✅ 生成 {OUT}：{len(blocks)} 条路线")


if __name__ == "__main__":
    main()
